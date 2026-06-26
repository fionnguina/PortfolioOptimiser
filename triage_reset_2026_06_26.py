"""One-off triage script (2026-06-26).

Why: the engine's _update_lots_after_trades was writing the *recommended*
trade plan to the Lots sheet every run instead of waiting for confirmed
fills. After many runs this inflated SMH lots to ~3.4M units, which the
live TLH pass then 'harvested' into a -1.7M / +1.7M trade plan worth
$6.3B in turnover.

IBKR has also reset the user's paper account to a $250k cap. Actual
positions per user: BBUS=531, BEAR=21590, HBRD=2034, $57,076 AUD cash.

This script:
  1. Backs up Stock Analysis.xlsm (already done in bash).
  2. Zeroes Holdings except the three real positions.
  3. Wipes Lots and seeds three rows (today, current AUD last px) so the
     CGT/TLH machinery has a baseline cost basis.
  4. Updates portfolio_state.json so portfolio_value matches reality.

Run once. Not idempotent in spirit (would re-truncate Lots if rerun).
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

APP_DIR = Path(__file__).resolve().parent
XL = APP_DIR / "Stock Analysis.xlsm"
STATE = APP_DIR / "portfolio_state.json"
SEED = APP_DIR / "lots_seed.json"

ACTUAL_POSITIONS = {
    "BBUS.AX": 531,
    "BEAR.AX": 21590,
    "HBRD.AX": 2034,
}
CASH_AUD = 57076.0
TODAY = datetime(2026, 6, 26)


def main() -> None:
    print(f"[triage] opening {XL}")
    wb = openpyxl.load_workbook(XL, keep_vba=True)

    # --- Holdings: zero everything except the three real positions -----------
    hold = wb["Holdings"]
    header = [c.value for c in hold[1]]
    sec_col = header.index("Security") + 1
    units_col = header.index("Units") + 1
    px_col = header.index("Last Price") + 1
    fx_col = header.index("FX to AUD") + 1
    mv_col = header.index("Market Value") + 1
    wt_col = header.index("Weight") + 1

    total_mv = 0.0
    px_for_lot: dict[str, float] = {}
    fx_for_lot: dict[str, float] = {}

    for r in range(2, hold.max_row + 1):
        sec = hold.cell(r, sec_col).value
        if sec is None:
            continue
        sec = str(sec).strip()
        new_units = ACTUAL_POSITIONS.get(sec, 0)
        hold.cell(r, units_col).value = new_units
        try:
            px = float(hold.cell(r, px_col).value or 0)
        except Exception:
            px = 0.0
        try:
            fx = float(hold.cell(r, fx_col).value or 1)
        except Exception:
            fx = 1.0
        mv_aud = new_units * px * fx
        hold.cell(r, mv_col).value = mv_aud
        total_mv += mv_aud
        if sec in ACTUAL_POSITIONS:
            px_for_lot[sec] = px
            fx_for_lot[sec] = fx

    # Second pass for weights (need total_mv first)
    if total_mv > 0:
        for r in range(2, hold.max_row + 1):
            try:
                mv = float(hold.cell(r, mv_col).value or 0)
            except Exception:
                mv = 0.0
            hold.cell(r, wt_col).value = mv / total_mv

    print(f"[triage] Holdings: invested MV = ${total_mv:,.0f} AUD")

    # --- Lots: wipe & seed three rows ----------------------------------------
    lots = wb["Lots"]
    # Clear all rows from row 2 onward
    if lots.max_row >= 2:
        lots.delete_rows(2, lots.max_row - 1)

    headers = ["Security", "AcqDate", "Units", "CostBaseAUD"]
    for i, h in enumerate(headers, start=1):
        lots.cell(1, i).value = h

    row = 2
    seed_payload = []
    for sec, qty in ACTUAL_POSITIONS.items():
        px = px_for_lot.get(sec, 0.0)
        fx = fx_for_lot.get(sec, 1.0)
        cost_per_unit_aud = px * fx
        lots.cell(row, 1).value = sec
        lots.cell(row, 2).value = TODAY
        lots.cell(row, 3).value = qty
        lots.cell(row, 4).value = cost_per_unit_aud
        seed_payload.append({
            "Security": sec,
            "AcqDate": TODAY.isoformat(),
            "Units": qty,
            "CostBaseAUD": cost_per_unit_aud,
        })
        print(f"[triage] Lot seeded: {sec:10s} qty={qty:>7,}  cost/unit AUD={cost_per_unit_aud:.4f}")
        row += 1

    wb.save(XL)
    print(f"[triage] saved {XL}")

    # Persist seed for the engine's fills-log rebuilder. Without this,
    # the rebuilder would wipe Lots on the next run (because the fills
    # log currently has zero filled rows).
    SEED.write_text(json.dumps(seed_payload, indent=2))
    print(f"[triage] wrote {SEED.name} ({len(seed_payload)} seed lots)")

    # --- portfolio_state.json -------------------------------------------------
    total_portfolio = total_mv + CASH_AUD
    state = {
        "portfolio_value": total_portfolio,
        "net_invested": total_mv,
    }
    STATE.write_text(json.dumps(state, indent=2))
    print(f"[triage] portfolio_state.json: portfolio_value=${total_portfolio:,.0f}  net_invested=${total_mv:,.0f}")

    print("[triage] done.")


if __name__ == "__main__":
    main()
